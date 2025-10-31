import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import requests
import urllib3
import json
from requests.auth import HTTPBasicAuth
import yfinance as yf
from datetime import datetime, timedelta
from curl_cffi import requests as curl_requests
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Disable SSL verification warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Page configuration
st.set_page_config(
    page_title="PCHP Analytics Dashboard",
    page_icon="📊",
    layout="wide"
)

st.title("📊 PCHP Analytics Dashboard")

# Session state for data
if 'df' not in st.session_state:
    st.session_state.df = None
    st.session_state.data_loaded = False

uri = st.secrets.db_credentials.url

# Auto-fetch data on first load
if not st.session_state.data_loaded:
    with st.spinner("Fetching data from API..."):
        try:
            base_url = uri
            auth = None
            params = {
                "BOName": "DASHBOARD FO",
                "Fields": "*",
                "Filter": "",
                "OrderBy": ""
            }
            
            response = requests.get(base_url, params=params, auth=auth, verify=False)
            
            if response.status_code == 200:
                content = response.content.decode("utf-8-sig")
                data = json.loads(content)
                
                if "OutputData" in data:
                    df = pd.DataFrame(data["OutputData"])
                    original_count = len(df)
                    
                    # Clean data
                    df = df[pd.to_numeric(df["TradeNumber"], errors="coerce").gt(0)]
                    cleaned_count = len(df)
                    removed_count = original_count - cleaned_count
                    
                    # Fix Position_Quantity
                    df["Position_Quantity"] = df.apply(
                        lambda row: row["CashflowUSD"] / row["NetPremium"] 
                                    if pd.isna(row["Position_Quantity"]) and row["NetPremium"] not in [0, None] 
                                    else row["Position_Quantity"],
                        axis=1
                    )
                    
                    st.session_state.df = df
                    st.session_state.data_loaded = True
                    st.success(f"✅ Data loaded! {cleaned_count} rows ({removed_count} removed)")
        except Exception as e:
            st.error(f"❌ Error loading data: {str(e)}")
            st.session_state.data_loaded = True

# Sidebar
st.sidebar.header("🔄 Data Management")

# Refresh data button
if st.sidebar.button("🔄 Refresh Data"):
    st.session_state.data_loaded = False
    st.rerun()

# Portfolio selector in sidebar
if st.session_state.df is not None:
    df = st.session_state.df
    st.sidebar.markdown("---")
    st.sidebar.header("📁 Portfolio Selection")
    portfolios = sorted(df["PortfolioName"].unique(), reverse=True)  # Sort descending to get latest first
    selected_portfolio = st.sidebar.selectbox("Select Portfolio", portfolios, index=0)  # Default to first (latest)
else:
    selected_portfolio = None

# Helper function
def format_volume(val):
    if abs(val) >= 1_000_000:
        return f"{val/1_000_000:.2f}M"
    elif abs(val) >= 1_000:
        return f"{val/1_000:.2f}k"
    else:
        return f"{val:.0f}"

def format_excel_worksheet(worksheet, df):
    """Format Excel worksheet with proper styling"""
    from openpyxl.styles import numbers
    
    # Define styles
    header_fill = PatternFill(start_color="00A19B", end_color="00A19B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_side = Side(style="thin", color="000000")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Define columns that should NOT use accounting format
    no_accounting_cols = ['TradeNumber', 'Position_Quantity', 'TradeDate', 'Settlement_DeliveryDate', 
                          'StartFixDate', 'EndFixDate', 'Strip', 'BuyOrSell', 'DealerID',
                          'OptionTypeLabel', 'StripTypeLabel', 'Portfolio']
    
    # Define date columns
    date_cols = ['TradeDate', 'Settlement_DeliveryDate', 'StartFixDate', 'EndFixDate']
    
    # Define monthly quantity columns (YYYY-MM format without USD)
    monthly_qty_cols = [col for col in df.columns if '-' in str(col) and 'USD' not in str(col)]
    
    # Format headers
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
        
        # Auto-adjust column width
        column_letter = get_column_letter(col_num)
        max_length = len(str(column_title))
        for row_num in range(2, len(df) + 2):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Format data cells
    for row_num in range(2, len(df) + 2):
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.alignment = center_alignment
            cell.border = border
            
            # Apply number formatting based on column type
            if column_title in date_cols:
                # Short date format (mm/dd/yyyy)
                cell.number_format = numbers.FORMAT_DATE_XLSX14
            elif column_title in no_accounting_cols or column_title in monthly_qty_cols:
                # No special formatting for these columns
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
            else:
                # Apply accounting format for all other numeric columns
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    
    return worksheet

def load_brent_data(start_date, end_date):
    """Load Brent Crude Futures data with fallback tickers using yfinance + curl_cffi session."""
    try:
        # Chrome-impersonated session fixes SSL / Yahoo blocking
        session = curl_requests.Session(impersonate="chrome", verify=False)
        
        # Common Yahoo tickers for Brent futures
        tickers = ["BZ=F"]
        
        data = pd.DataFrame()
        for ticker in tickers:
            temp = yf.download(ticker, start=start_date, end=end_date, session=session, progress=False)
            if not temp.empty:
                data = temp.copy()
                data["Ticker"] = ticker
                break
        
        if data.empty:
            return pd.DataFrame()
        
        # Handle MultiIndex columns (sometimes appears in yfinance)
        if isinstance(data.columns, pd.MultiIndex):
            data.columns = data.columns.get_level_values(0)
        
        data = data.reset_index()
        data["Date"] = pd.to_datetime(data["Date"])
        return data
    
    except Exception as e:
        st.error(f"Error loading Brent data: {str(e)}")
        return pd.DataFrame()

# Main content
if st.session_state.df is not None:
    df = st.session_state.df
    
    # Tabs for different views
    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "📋 Summary Table", 
        "📊 Dealer Volumes", 
        "📅 Monthly Analysis",
        "📉 Brent Price & Trades",
        "📄 MTM Analysis",
        "🎯 ITM Analysis",
        "📈 Raw Data", 
        "💾 Export"
    ])
    
    # Tab 1: Portfolio Summary Table
    with tab1:
        st.header("Portfolio Trade Summary")
        
        def portfolio_summary(group):
            num_trades = group["TradeNumber"].count()
            total_pos_qty = group.loc[group["BuyOrSell"] == 1, "Position_Quantity"].sum()
            abs_net_prem = group["AbsNetPremium"].sum()
            
            # Weighted Protection Level
            protection_leg = group.loc[group["BuyOrSell"] == 1]
            if not protection_leg.empty and protection_leg["Position_Quantity"].sum() != 0:
                prot_num = (protection_leg["Position_Quantity"] * protection_leg["StrikePrice1"]).sum()
                prot_den = protection_leg["Position_Quantity"].sum()
                weighted_protection = prot_num / prot_den
            else:
                weighted_protection = 0
            
            # Weighted Lower Level
            strike1 = pd.to_numeric(group["StrikePrice1"], errors="coerce").fillna(0)
            strike2 = pd.to_numeric(group["StrikePrice2"], errors="coerce").fillna(0)
            
            num_buy = (group.loc[group["BuyOrSell"] == 1, "Position_Quantity"] *
                       strike2.loc[group["BuyOrSell"] == 1]).sum()
            num_sell = (group.loc[group["BuyOrSell"] != 1, "Position_Quantity"] *
                        strike1.loc[group["BuyOrSell"] != 1]).sum()
            lower_num = num_buy + num_sell
            
            den_buy = group.loc[(group["BuyOrSell"] == 1) & (strike2 != 0), "Position_Quantity"].sum()
            den_sell = group.loc[group["BuyOrSell"] != 1, "Position_Quantity"].sum()
            lower_den = den_buy + den_sell
            
            weighted_lower = lower_num / lower_den if lower_den != 0 else 0
            
            return pd.Series({
                "Number_of_Trades": num_trades,
                "Total_Position_Quantity": total_pos_qty,
                "AbsNetPremium": abs_net_prem,
                "Weighted_Protection_Level": weighted_protection,
                "Weighted_Lower_Level": weighted_lower
            })
        
        summary = df.groupby("PortfolioName", sort=False).apply(portfolio_summary).reset_index()
        
        # Sort by portfolio name descending (latest/highest year first)
        summary = summary.sort_values("PortfolioName", ascending=False)
        
        # Format for display
        summary_display = summary.copy()
        summary_display["Number_of_Trades"] = summary_display["Number_of_Trades"].apply(lambda x: f"{x:,}")
        summary_display["Total_Position_Quantity"] = summary_display["Total_Position_Quantity"].apply(lambda x: f"{x:,.0f}")
        summary_display["AbsNetPremium"] = summary_display["AbsNetPremium"].apply(lambda x: f"${x:,.2f}")
        summary_display["Weighted_Protection_Level"] = summary_display["Weighted_Protection_Level"].apply(lambda x: f"{x:.2f}")
        summary_display["Weighted_Lower_Level"] = summary_display["Weighted_Lower_Level"].apply(lambda x: f"{x:.2f}")
        
        summary_display.columns = [
            "Portfolio Name",
            "Number of Trades",
            "Total Position Quantity",
            "Total Cost",
            "Weighted Protection Level",
            "Weighted Lower Level"
        ]
        
        st.dataframe(summary_display, use_container_width=True, hide_index=True)
        
        # Position Quantity by Portfolio and Acronym
        st.subheader("Position Quantity by Portfolio and Acronym")
        
        # Pivot table: rows = PortfolioName, columns = Acronym, values = sum(Position_Quantity)
        pivot_df = df.pivot_table(
            index="PortfolioName",
            columns="Acronym",
            values="Position_Quantity",
            aggfunc="sum",
            fill_value=0
        )
        
        # Sort columns alphabetically and rows by portfolio name descending
        pivot_df = pivot_df.reindex(sorted(pivot_df.columns), axis=1)
        pivot_df = pivot_df.sort_index(ascending=False)
        
        # Format all pivot table values
        formatted_pivot = pivot_df.applymap(lambda x: format_volume(x) if x != 0 else "0")
        
        # Reset index to make PortfolioName a column
        formatted_pivot = formatted_pivot.reset_index()
        
        st.dataframe(formatted_pivot, use_container_width=True, hide_index=True)
    
    # Tab 2: Dealer Volumes Chart
    with tab2:
        st.header(f"Dealer Volumes - {selected_portfolio}")
        
        if selected_portfolio:
            # Filter portfolio and BuyOrSell == 1
            portfolio_df = df[df["PortfolioName"] == selected_portfolio].copy()
            portfolio_df["BuyOrSell"] = pd.to_numeric(portfolio_df["BuyOrSell"], errors="coerce")
            portfolio_df = portfolio_df[portfolio_df["BuyOrSell"] == 1].copy()
            
            dealers = portfolio_df["DealerID"].unique()
            agg_df = portfolio_df.groupby(
                ["Acronym", "DealerID"], 
                as_index=False, 
                dropna=False
            )["Position_Quantity"].sum()
            
            full_index = pd.MultiIndex.from_product(
                [agg_df["Acronym"].unique(), dealers], 
                names=["Acronym", "DealerID"]
            )
            agg_df = agg_df.set_index(["Acronym", "DealerID"]).reindex(full_index, fill_value=0).reset_index()
            
            totals = agg_df.groupby("Acronym", as_index=False)["Position_Quantity"].sum()
            totals = totals.sort_values("Position_Quantity", ascending=False)
            sorted_acronyms = totals["Acronym"].tolist()
            
            # Create figure
            fig = make_subplots(
                rows=2, cols=1,
                shared_xaxes=True,
                vertical_spacing=0.15,
                subplot_titles=("Dealer Volumes by Acronym", "Total Volume by Acronym"),
                row_heights=[0.6, 0.4]
            )
            
            # Add dealer traces
            for dealer in dealers:
                dealer_data = agg_df[agg_df["DealerID"] == dealer].copy()
                dealer_data["Acronym"] = pd.Categorical(
                    dealer_data["Acronym"], 
                    categories=sorted_acronyms, 
                    ordered=True
                )
                dealer_data = dealer_data.sort_values("Acronym")
                
                fig.add_trace(
                    go.Bar(
                        x=dealer_data["Acronym"],
                        y=dealer_data["Position_Quantity"],
                        name=str(dealer),
                        text=[format_volume(v) if v != 0 else "" for v in dealer_data["Position_Quantity"]],
                        textposition="inside",
                        textfont=dict(size=10),
                        hovertemplate="<b>%{x}</b><br>Dealer: " + str(dealer) + "<br>Volume: %{y:,.0f}<extra></extra>"
                    ),
                    row=1, col=1
                )
            
            # Add total trace
            totals_sorted = totals.copy()
            totals_sorted["Acronym"] = pd.Categorical(
                totals_sorted["Acronym"], 
                categories=sorted_acronyms, 
                ordered=True
            )
            totals_sorted = totals_sorted.sort_values("Acronym")
            
            fig.add_trace(
                go.Bar(
                    x=totals_sorted["Acronym"],
                    y=totals_sorted["Position_Quantity"],
                    name="Total",
                    marker=dict(color="#2C3E50"),
                    text=[format_volume(v) if v != 0 else "" for v in totals_sorted["Position_Quantity"]],
                    textposition="inside",
                    textfont=dict(size=11, color="white"),
                    hovertemplate="<b>%{x}</b><br>Total Volume: %{y:,.0f}<extra></extra>"
                ),
                row=2, col=1
            )
            
            fig.update_layout(
                barmode="group",
                title=dict(
                    text=f"Dealer Volumes & Totals - {selected_portfolio}",
                    x=0.5,
                    xanchor="center"
                ),
                height=800,
                hovermode="x unified",
                legend=dict(
                    orientation="v",
                    yanchor="top",
                    y=0.99,
                    xanchor="left",
                    x=1.02
                ),
                xaxis=dict(title="Acronym"),
                xaxis2=dict(title="Acronym"),
                yaxis=dict(title="Volume"),
                yaxis2=dict(title="Volume")
            )
            
            st.plotly_chart(fig, use_container_width=True)
    
    # Tab 3: Monthly Analysis
    with tab3:
        st.header(f"Monthly Analysis - {selected_portfolio}")
        
        if selected_portfolio:
            # Filter portfolio and BuyOrSell == 1
            portfolio_df = df[df["PortfolioName"] == selected_portfolio].copy()
            portfolio_df["BuyOrSell"] = pd.to_numeric(portfolio_df["BuyOrSell"], errors="coerce")
            portfolio_df = portfolio_df[portfolio_df["BuyOrSell"] == 1].copy()
            
            portfolio_df["StartFixDate"] = pd.to_datetime(portfolio_df["StartFixDate"])
            portfolio_df["EndFixDate_MaturityDate"] = pd.to_datetime(portfolio_df["EndFixDate_MaturityDate"])
            
            # Prepare monthly breakdown
            monthly_rows = []
            for _, row in portfolio_df.iterrows():
                start = row["StartFixDate"].replace(day=1)
                end = row["EndFixDate_MaturityDate"].replace(day=1)
                months_diff = (end.year - start.year) * 12 + (end.month - start.month) + 1
                monthly_quantity = row["Position_Quantity"] / months_diff

                for m in range(months_diff):
                    month = start + pd.DateOffset(months=m)
                    monthly_rows.append({
                        "Month": month,
                        "DealerID": row["DealerID"],
                        "StrikePrice1": row["StrikePrice1"],
                        "Acronym": row["Acronym"],
                        "Monthly_Position_Quantity": monthly_quantity
                    })

            monthly_df = pd.DataFrame(monthly_rows)
            
            # Chart 1: DealerID breakdown
            st.subheader("Monthly Position by DealerID")
            agg_dealer = monthly_df.groupby(["Month", "DealerID"], as_index=False)["Monthly_Position_Quantity"].sum()
            
            fig_dealer = go.Figure()
            for dealer in sorted(agg_dealer["DealerID"].unique()):
                data = agg_dealer[agg_dealer["DealerID"] == dealer]
                fig_dealer.add_trace(go.Bar(
                    x=data["Month"],
                    y=data["Monthly_Position_Quantity"],
                    name=str(dealer),
                    text=[format_volume(v) for v in data["Monthly_Position_Quantity"]],
                    textposition="inside",
                    textfont=dict(size=10),
                    hovertemplate="<b>%{x|%b %Y}</b><br>DealerID: " + str(dealer) +
                                  "<br>Volume: %{y:,.0f}<extra></extra>"
                ))

            fig_dealer.update_layout(
                barmode="stack",
                height=600,
                hovermode="x unified",
                legend=dict(orientation="v", yanchor="top", y=0.99, xanchor="left", x=1.02),
                xaxis=dict(title="Month"),
                yaxis=dict(title="Volume")
            )
            st.plotly_chart(fig_dealer, use_container_width=True)
            
            # Chart 2: StrikePrice1 breakdown (already filtered for BuyOrSell=1)
            st.subheader("Monthly Position by StrikePrice1")
            
            monthly_rows_strike = []
            for _, row in portfolio_df.iterrows():
                start = row["StartFixDate"].replace(day=1)
                end = row["EndFixDate_MaturityDate"].replace(day=1)
                months_diff = (end.year - start.year) * 12 + (end.month - start.month) + 1
                monthly_quantity = row["Position_Quantity"] / months_diff

                for m in range(months_diff):
                    month = start + pd.DateOffset(months=m)
                    monthly_rows_strike.append({
                        "Month": month,
                        "StrikePrice1": row["StrikePrice1"],
                        "Monthly_Position_Quantity": monthly_quantity
                    })

            monthly_df_strike = pd.DataFrame(monthly_rows_strike)
            agg_strike = monthly_df_strike.groupby(["Month", "StrikePrice1"], as_index=False)["Monthly_Position_Quantity"].sum()

            fig_strike = go.Figure()
            for strike in sorted(agg_strike["StrikePrice1"].unique()):
                data = agg_strike[agg_strike["StrikePrice1"] == strike]
                fig_strike.add_trace(go.Bar(
                    x=data["Month"],
                    y=data["Monthly_Position_Quantity"],
                    name=str(strike),
                    text=[format_volume(v) for v in data["Monthly_Position_Quantity"]],
                    textposition="inside",
                    textfont=dict(size=10),
                    hovertemplate="<b>%{x|%b %Y}</b><br>StrikePrice1: " + str(strike) +
                                  "<br>Volume: %{y:,.0f}<extra></extra>"
                ))

            fig_strike.update_layout(
                barmode="stack",
                height=600,
                hovermode="x unified",
                legend=dict(orientation="v", yanchor="top", y=0.99, xanchor="left", x=1.02),
                xaxis=dict(title="Month"),
                yaxis=dict(title="Volume")
            )
            st.plotly_chart(fig_strike, use_container_width=True)
            
            # Chart 3: Acronym breakdown with total
            st.subheader("Monthly Position by Acronym")
            agg_acronym = monthly_df.groupby(["Month", "Acronym"], as_index=False)["Monthly_Position_Quantity"].sum()
            total_acronym = monthly_df.groupby("Month", as_index=False)["Monthly_Position_Quantity"].sum()

            fig_acronym = make_subplots(
                rows=2, cols=1,
                shared_xaxes=True,
                vertical_spacing=0.1,
                subplot_titles=("Monthly Position Quantity by Acronym", "Total Monthly Position Quantity")
            )

            for acronym in sorted(agg_acronym["Acronym"].unique()):
                data = agg_acronym[agg_acronym["Acronym"] == acronym]
                fig_acronym.add_trace(go.Bar(
                    x=data["Month"],
                    y=data["Monthly_Position_Quantity"],
                    name=str(acronym),
                    text=[format_volume(v) for v in data["Monthly_Position_Quantity"]],
                    textposition="inside",
                    textfont=dict(size=10),
                    hovertemplate="<b>%{x|%b %Y}</b><br>Acronym: " + str(acronym) +
                                  "<br>Volume: %{y:,.0f}<extra></extra>"
                ), row=1, col=1)

            fig_acronym.add_trace(go.Bar(
                x=total_acronym["Month"],
                y=total_acronym["Monthly_Position_Quantity"],
                name="Total",
                marker_color="#2C3E50",
                text=[format_volume(v) for v in total_acronym["Monthly_Position_Quantity"]],
                textposition="inside",
                textfont=dict(size=11, color="white"),
                hovertemplate="<b>%{x|%b %Y}</b><br>Total Volume: %{y:,.0f}<extra></extra>"
            ), row=2, col=1)

            fig_acronym.update_layout(
                barmode="stack",
                height=800,
                hovermode="x unified",
                legend=dict(orientation="v", yanchor="top", y=0.99, xanchor="left", x=1.02),
                xaxis=dict(title="Month"),
                yaxis=dict(title="Volume"),
                yaxis2=dict(title="Volume")
            )
            st.plotly_chart(fig_acronym, use_container_width=True)
    
    # Tab 4: Brent Price & Trades
    with tab4:
        st.header(f"Brent Price & Trade Dates - {selected_portfolio}")
        
        if selected_portfolio:
            # Filter by selected portfolio
            portfolio_df = df[df["PortfolioName"] == selected_portfolio].copy()
            
            # Convert TradeDate to datetime
            portfolio_df["TradeDate"] = pd.to_datetime(portfolio_df["TradeDate"], errors="coerce")
            trade_dates = portfolio_df["TradeDate"].dropna().sort_values().unique()
            
            if len(trade_dates) > 0:
                # Get date range with buffer
                start_date = (pd.to_datetime(trade_dates.min()) - timedelta(days=30)).strftime('%Y-%m-%d')
                end_date = (pd.to_datetime(trade_dates.max()) + timedelta(days=30)).strftime('%Y-%m-%d')
                
                # Fetch Brent crude oil data
                with st.spinner("Fetching Brent crude oil prices..."):
                    brent = load_brent_data(start_date, end_date)
                    
                    if not brent.empty:
                        # Prepare monthly breakdown for trade date analysis
                        portfolio_df["StartFixDate"] = pd.to_datetime(portfolio_df["StartFixDate"])
                        portfolio_df["EndFixDate_MaturityDate"] = pd.to_datetime(portfolio_df["EndFixDate_MaturityDate"])
                        
                        # Create monthly breakdown per trade
                        trade_monthly_breakdown = []
                        for _, row in portfolio_df.iterrows():
                            start = row["StartFixDate"].replace(day=1)
                            end = row["EndFixDate_MaturityDate"].replace(day=1)
                            months_diff = (end.year - start.year) * 12 + (end.month - start.month) + 1
                            monthly_quantity = row["Position_Quantity"] / months_diff
                            
                            for m in range(months_diff):
                                month = start + pd.DateOffset(months=m)
                                trade_monthly_breakdown.append({
                                    "TradeDate": row["TradeDate"],
                                    "Month": month,
                                    "Monthly_Position_Quantity": monthly_quantity
                                })
                        
                        trade_monthly_df = pd.DataFrame(trade_monthly_breakdown)
                        trade_monthly_df["TradeDate"] = pd.to_datetime(trade_monthly_df["TradeDate"])
                        
                        # Group by TradeDate and Month to get quantities
                        trade_month_summary = trade_monthly_df.groupby(["TradeDate", "Month"])["Monthly_Position_Quantity"].sum().reset_index()
                        
                        # Calculate position quantities per trade date (total)
                        trade_date_quantities = portfolio_df.groupby("TradeDate")["Position_Quantity"].sum().reset_index()
                        trade_date_quantities["TradeDate"] = pd.to_datetime(trade_date_quantities["TradeDate"])
                        
                        # Create figure
                        fig_brent = go.Figure()
                        
                        # Add Brent price line
                        fig_brent.add_trace(go.Scatter(
                            x=brent["Date"],
                            y=brent["Close"],
                            mode="lines",
                            name="Brent Price",
                            line=dict(color="#2E86AB", width=2),
                            hovertemplate="<b>%{x|%b %d, %Y}</b><br>Price: $%{y:.2f}<extra></extra>"
                        ))
                        
                        # Add markers for trade dates with monthly breakdown
                        trade_prices = []
                        trade_quantities = []
                        trade_monthly_info = []
                        valid_trade_dates = []
                        
                        for _, row in trade_date_quantities.iterrows():
                            trade_date = row["TradeDate"]
                            quantity = row["Position_Quantity"]
                            
                            # Find closest Brent price for this trade date
                            closest_idx = brent[brent["Date"] <= trade_date]["Date"].idxmax() if not brent[brent["Date"] <= trade_date].empty else None
                            if closest_idx is not None:
                                trade_prices.append(brent.loc[closest_idx, "Close"])
                                trade_quantities.append(quantity)
                                valid_trade_dates.append(trade_date)
                                
                                # Get monthly breakdown for this trade date
                                monthly_data = trade_month_summary[trade_month_summary["TradeDate"] == trade_date]
                                monthly_breakdown = "<br>".join([
                                    f"  {row['Month'].strftime('%b %Y')}: {format_volume(row['Monthly_Position_Quantity'])}"
                                    for _, row in monthly_data.iterrows()
                                ])
                                trade_monthly_info.append(monthly_breakdown)
                        
                        fig_brent.add_trace(go.Scatter(
                            x=valid_trade_dates,
                            y=trade_prices,
                            mode="markers",
                            name="Trade Dates",
                            marker=dict(
                                size=10,
                                color="#A23B72",
                                symbol="diamond",
                                line=dict(color="white", width=2)
                            ),
                            text=[format_volume(q) for q in trade_quantities],
                            customdata=[[q, info] for q, info in zip(trade_quantities, trade_monthly_info)],
                            hovertemplate="<b>Trade Date: %{x|%b %d, %Y}</b><br>" +
                                        "Brent Price: $%{y:.2f}<br>" +
                                        "Total Position: %{customdata[0]:,.0f}<br>" +
                                        "<b>Monthly Breakdown:</b><br>%{customdata[1]}<extra></extra>"
                        ))
                        
                        fig_brent.update_layout(
                            title=f"Brent Crude Oil Price with Trade Dates - {selected_portfolio}",
                            xaxis_title="Date",
                            yaxis_title="Price (USD per Barrel)",
                            height=600,
                            hovermode="x unified",
                            showlegend=True,
                            legend=dict(
                                orientation="h",
                                yanchor="bottom",
                                y=1.02,
                                xanchor="right",
                                x=1
                            )
                        )
                        
                        st.plotly_chart(fig_brent, use_container_width=True)
                    else:
                        st.warning("Unable to fetch Brent price data")
            else:
                st.warning("No valid trade dates found for this portfolio")
    
    # Tab 5: Trade Details
    with tab5:
        st.header(f"Trade Details - {selected_portfolio}")
        
        if selected_portfolio:
            # Option to include/exclude expired volumes
            include_expired = st.checkbox(
                "Include expired volumes (past months)",
                value=False,
                help="If unchecked, monthly quantities for past months will be set to zero"
            )
            
            # Filter by selected portfolio
            trade_details_df = df[df["PortfolioName"] == selected_portfolio].copy()
            
            # Select specific columns
            columns_to_display = [
                "TradeNumber",
                "PortfolioName",
                "TradeDate",
                "Settlement_DeliveryDate",
                "StrikePrice1",
                "StrikePrice2",
                "NetPremium",
                "AbsNetPremium",
                "BuyOrSell",
                "Position_Quantity",
                "DealerID",
                "OptionTypeLabel",
                "StripTypeLabel",
                "StartFixDate",
                "EndFixDate_MaturityDate"
            ]
            
            # Check which columns exist in the dataframe
            available_columns = [col for col in columns_to_display if col in trade_details_df.columns]
            
            # Select only available columns
            trade_details_display = trade_details_df[available_columns].copy()
            
            # Calculate monthly breakdown for each trade
            trade_details_display["StartFixDate"] = pd.to_datetime(trade_details_display["StartFixDate"])
            trade_details_display["EndFixDate_MaturityDate"] = pd.to_datetime(trade_details_display["EndFixDate_MaturityDate"])
            
            # Add Strip column
            trade_details_display["Strip"] = trade_details_display.apply(
                lambda row: (row["EndFixDate_MaturityDate"].replace(day=1) - row["StartFixDate"].replace(day=1)).days > 31,
                axis=1
            )
            
            # Get current month for comparison
            current_month = pd.Timestamp.now().replace(day=1)
            
            # Create a list to store all month columns
            all_months = set()
            trade_monthly_data = {}
            
            for idx, row in trade_details_display.iterrows():
                start = row["StartFixDate"].replace(day=1)
                end = row["EndFixDate_MaturityDate"].replace(day=1)
                months_diff = (end.year - start.year) * 12 + (end.month - start.month) + 1
                monthly_quantity = row["Position_Quantity"] / months_diff
                
                trade_monthly_data[idx] = {}
                for m in range(months_diff):
                    month = start + pd.DateOffset(months=m)
                    month_str = month.strftime('%Y-%m')
                    all_months.add(month_str)
                    
                    # If not including expired volumes and month is in the past, set to 0
                    if not include_expired and month < current_month:
                        trade_monthly_data[idx][month_str] = 0
                    else:
                        trade_monthly_data[idx][month_str] = monthly_quantity
            
            # Sort months chronologically
            sorted_months = sorted(list(all_months))
            
            # Add monthly columns to the dataframe
            for month in sorted_months:
                trade_details_display[month] = trade_details_display.index.map(
                    lambda idx: trade_monthly_data.get(idx, {}).get(month, 0)
                )
            
            # Rename columns for better display
            column_rename = {
                "PortfolioName": "Portfolio",
                "EndFixDate_MaturityDate": "EndFixDate"
            }
            trade_details_display = trade_details_display.rename(columns=column_rename)
            
            # Reorder columns to place Strip after EndFixDate
            cols = trade_details_display.columns.tolist()
            if "Strip" in cols and "EndFixDate" in cols:
                # Remove Strip from its current position
                cols.remove("Strip")
                # Insert Strip right after EndFixDate
                endfix_idx = cols.index("EndFixDate")
                cols.insert(endfix_idx + 1, "Strip")
                trade_details_display = trade_details_display[cols]
            
            # Sort by TradeNumber ascending (most recent first)
            if "TradeNumber" in trade_details_display.columns:
                trade_details_display = trade_details_display.sort_values("TradeNumber", ascending=True)
            
            # Format numeric columns
            numeric_columns = ["StrikePrice1", "StrikePrice2", "NetPremium", "Position_Quantity", "AbsNetPremium"] + sorted_months
            for col in numeric_columns:
                if col in trade_details_display.columns:
                    trade_details_display[col] = trade_details_display[col].apply(
                        lambda x: f"{x:,.2f}" if pd.notna(x) and x != 0 else "0"
                    )
            
            st.dataframe(trade_details_display, use_container_width=True, height=600)
            st.info(f"Total trades for {selected_portfolio}: {len(trade_details_display):,} | Monthly columns: {len(sorted_months)}")
            
            st.markdown("---")
            
            # Create two columns for template download and upload
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("Download Strike Price Template")
                
                if sorted_months:
                    # Get unique StrikePrice1 and StrikePrice2 values
                    unique_strike1 = sorted(trade_details_df["StrikePrice1"].dropna().unique())
                    unique_strike2 = sorted(trade_details_df["StrikePrice2"].dropna().unique())
                    
                    # Combine and get unique strikes
                    all_strikes = sorted(set(list(unique_strike1) + list(unique_strike2)))
                    
                    # Filter months based on include_expired checkbox
                    if include_expired:
                        template_months = sorted_months
                    else:
                        # Only include current month and future months
                        template_months = [m for m in sorted_months if pd.to_datetime(m + "-01") >= current_month]
                    
                    if template_months:
                        # Create empty template dataframe
                        template_df = pd.DataFrame(index=all_strikes, columns=template_months)
                        template_df.index.name = "Strike Price"
                        
                        # Fill with empty strings
                        template_df = template_df.fillna("")
                        
                        # Convert to Excel with formatting
                        output = BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            template_df.to_excel(writer, sheet_name='Strike Prices')
                            
                            # Format the worksheet
                            worksheet = writer.sheets['Strike Prices']
                            format_excel_worksheet(worksheet, template_df.reset_index())
                        
                        output.seek(0)

                        today = datetime.now().strftime("%Y-%m-%d")
                        
                        st.download_button(
                            label="📥 Download Template",
                            data=output,
                            file_name=f"{today}_strike_price_template_{selected_portfolio}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.info(f"📊 {len(all_strikes)} strikes × {len(template_months)} months")
                    else:
                        st.warning("No future months available for template")
                else:
                    st.warning("No monthly data available")
            
            with col2:
                st.subheader("Upload Filled Template")
                
                uploaded_template = st.file_uploader(
                    "Upload your filled Excel template",
                    type=['xlsx', 'xls'],
                    key=f"upload_template_{selected_portfolio}"
                )
                
                if uploaded_template is not None:
                    try:
                        # Read the uploaded Excel file
                        uploaded_df = pd.read_excel(uploaded_template, sheet_name='Strike Prices', index_col=0)
                        st.success("✅ File uploaded successfully!")
                        st.info(f"📊 {len(uploaded_df)} strikes × {len(uploaded_df.columns)} months")
                    except Exception as e:
                        st.error(f"Error reading file: {str(e)}")
                        uploaded_df = None
                else:
                    uploaded_df = None
            
            # Display uploaded data if available
            if uploaded_df is not None:
                st.markdown("---")
                st.subheader("Uploaded Strike Price Data")
                
                # Format the dataframe for better display
                display_df = uploaded_df.copy()
                
                # Replace NaN with empty string for display
                display_df = display_df.fillna("")
                
                st.dataframe(display_df, use_container_width=True, height=400)
                
                # Option to download the uploaded data as CSV
                csv = display_df.to_csv()
                st.download_button(
                    label="📥 Download as CSV",
                    data=csv,
                    file_name=f"strike_price_data_{selected_portfolio}.csv",
                    mime="text/csv"
                )
                
                # Calculate combined trade details with pricing
                st.markdown("---")
                st.subheader("Trade Details with Pricing Calculations")
                
                # Create a copy of the original trade details for calculation
                combined_df = trade_details_df[available_columns].copy()
                combined_df["StartFixDate"] = pd.to_datetime(combined_df["StartFixDate"])
                combined_df["EndFixDate_MaturityDate"] = pd.to_datetime(combined_df["EndFixDate_MaturityDate"])
                
                # Add Strip column
                combined_df["Strip"] = combined_df.apply(
                    lambda row: (row["EndFixDate_MaturityDate"].replace(day=1) - row["StartFixDate"].replace(day=1)).days > 31,
                    axis=1
                )
                
                # Calculate if each trade is expired (all months are in the past)
                combined_df["Is_Expired"] = combined_df.apply(
                    lambda row: row["EndFixDate_MaturityDate"].replace(day=1) < current_month, 
                    axis=1
                )
                
                # Add monthly quantity columns first
                for idx, row in combined_df.iterrows():
                    start = row["StartFixDate"].replace(day=1)
                    end = row["EndFixDate_MaturityDate"].replace(day=1)
                    months_diff = (end.year - start.year) * 12 + (end.month - start.month) + 1
                    monthly_quantity = row["Position_Quantity"] / months_diff
                    
                    for m in range(months_diff):
                        month = start + pd.DateOffset(months=m)
                        month_str = month.strftime('%Y-%m')
                        if month_str not in combined_df.columns:
                            combined_df[month_str] = 0.0
                        
                        # If not including expired volumes and month is in the past, set to 0
                        if not include_expired and month < current_month:
                            combined_df.at[idx, month_str] = 0
                        else:
                            combined_df.at[idx, month_str] = monthly_quantity
                
                # Calculate USD columns for each month
                for month_col in sorted_months:
                    if month_col in uploaded_df.columns:
                        usd_col_name = f"{month_col} (USD)"
                        combined_df[usd_col_name] = 0.0
                        
                        for idx, row in combined_df.iterrows():
                            if month_col in combined_df.columns and combined_df.at[idx, month_col] != 0:
                                monthly_qty = combined_df.at[idx, month_col]
                                strike1 = row["StrikePrice1"]
                                strike2 = row["StrikePrice2"]
                                buy_or_sell = row["BuyOrSell"]
                                
                                # Get price for Strike1 from uploaded template
                                price1 = 0
                                try:
                                    if pd.notna(strike1) and strike1 in uploaded_df.index:
                                        price1_val = uploaded_df.at[strike1, month_col]
                                        price1 = float(price1_val) if pd.notna(price1_val) and price1_val != "" else 0
                                except:
                                    price1 = 0
                                
                                # Calculate Strike1 component: Quantity × Price (NOT Strike × Quantity × Price)
                                strike1_value = monthly_qty * price1
                                
                                # Get price for Strike2 from uploaded template (if exists)
                                price2 = 0
                                strike2_value = 0
                                if pd.notna(strike2) and strike2 != 0:
                                    try:
                                        if strike2 in uploaded_df.index:
                                            price2_val = uploaded_df.at[strike2, month_col]
                                            price2 = float(price2_val) if pd.notna(price2_val) and price2_val != "" else 0
                                    except:
                                        price2 = 0
                                    
                                    # Calculate Strike2 component: Quantity × Price (NOT Strike × Quantity × Price)
                                    strike2_value = monthly_qty * price2
                                
                                # Calculate final USD value
                                # If BuyOrSell is None/NaN/0, negate the value
                                if pd.isna(buy_or_sell) or buy_or_sell == 0 or buy_or_sell == "":
                                    usd_value = -(strike1_value - strike2_value)
                                else:
                                    usd_value = strike1_value - strike2_value
                                
                                combined_df.at[idx, usd_col_name] = usd_value
                
                # Rename columns for display
                column_rename = {
                    "PortfolioName": "Portfolio",
                    "EndFixDate_MaturityDate": "EndFixDate"
                }
                combined_df = combined_df.rename(columns=column_rename)
                
                # Sort by TradeNumber ascending
                if "TradeNumber" in combined_df.columns:
                    combined_df = combined_df.sort_values("TradeNumber", ascending=True)
                
                # Reorder columns: base columns, then monthly quantities, then USD columns
                base_cols = ["TradeNumber", "Portfolio", "TradeDate", "Settlement_DeliveryDate", 
                            "StrikePrice1", "StrikePrice2", "NetPremium", "AbsNetPremium", "BuyOrSell",
                            "Position_Quantity", "DealerID", "OptionTypeLabel", "StripTypeLabel",
                            "StartFixDate", "EndFixDate", "Strip"]
                
                # Get columns that exist in the dataframe
                existing_base_cols = [col for col in base_cols if col in combined_df.columns]
                
                # Get monthly quantity columns (sorted chronologically)
                monthly_qty_cols = sorted([col for col in combined_df.columns if col in sorted_months])
                
                # Get USD columns (sorted chronologically by extracting the month part)
                usd_cols = [col for col in combined_df.columns if "(USD)" in col]
                usd_cols_sorted = sorted(usd_cols, key=lambda x: x.replace(" (USD)", ""))
                
                # Calculate Current Outstanding Value (sum of all USD columns)
                combined_df["Current Outstanding Value, USD"] = combined_df[usd_cols_sorted].sum(axis=1)
                
                # Calculate Current Premium/Barrel
                combined_df["Current Premium/Barrel"] = combined_df.apply(
                    lambda row: row["Current Outstanding Value, USD"] / row["Position_Quantity"] 
                    if row["Position_Quantity"] != 0 else 0,
                    axis=1
                )
                
                # Calculate metrics for Mark to Market
                total_inception_value = combined_df["AbsNetPremium"].sum()
                outstanding_inception_value = combined_df[~combined_df["Is_Expired"]]["AbsNetPremium"].sum()
                current_outstanding_value = combined_df["Current Outstanding Value, USD"].sum()
                mtm_movements = current_outstanding_value - outstanding_inception_value
                mtm_percentage = (mtm_movements / outstanding_inception_value * 100) if outstanding_inception_value != 0 else 0
                
                # Display Mark to Market metrics
                st.subheader("Mark to Market Result")
                
                col_mtm1, col_mtm2, col_mtm3, col_mtm4 = st.columns(4)
                
                with col_mtm1:
                    st.metric(
                        label="Total Inception Value, USD",
                        value=f"${total_inception_value:,.2f}"
                    )
                
                with col_mtm2:
                    st.metric(
                        label="Outstanding Inception Value, USD",
                        value=f"${outstanding_inception_value:,.2f}"
                    )
                
                with col_mtm3:
                    st.metric(
                        label="Current Outstanding Value, USD",
                        value=f"${current_outstanding_value:,.2f}"
                    )
                
                with col_mtm4:
                    st.metric(
                        label="MTM Movements, USD",
                        value=f"${mtm_movements:,.2f}",
                        delta=f"{mtm_percentage:+.2f}%"
                    )
                
                st.markdown("---")
                
                # Combine in order: base columns, monthly quantities, USD values, outstanding value, premium per barrel
                ordered_columns = existing_base_cols + monthly_qty_cols + usd_cols_sorted + ["Current Outstanding Value, USD", "Current Premium/Barrel"]
                
                # Reorder the dataframe (exclude Is_Expired from display)
                combined_df = combined_df[ordered_columns]
                
                # Format numeric columns for display
                display_combined = combined_df.copy()
                
                # Format base numeric columns
                base_numeric = ["StrikePrice1", "StrikePrice2", "NetPremium", "Position_Quantity", "AbsNetPremium"]
                for col in base_numeric:
                    if col in display_combined.columns:
                        display_combined[col] = display_combined[col].apply(
                            lambda x: f"{x:,.2f}" if pd.notna(x) else ""
                        )
                
                # Format monthly quantity columns
                for month_col in sorted_months:
                    if month_col in display_combined.columns:
                        display_combined[month_col] = display_combined[month_col].apply(
                            lambda x: f"{x:,.2f}" if pd.notna(x) and x != 0 else "0"
                        )
                
                # Format USD columns
                usd_columns = [col for col in display_combined.columns if "(USD)" in col or col == "Current Outstanding Value, USD"]
                for col in usd_columns:
                    display_combined[col] = display_combined[col].apply(
                        lambda x: f"${x:,.2f}" if pd.notna(x) else "$0.00"
                    )
                
                # Format Current Premium/Barrel with 3 decimal places
                if "Current Premium/Barrel" in display_combined.columns:
                    display_combined["Current Premium/Barrel"] = display_combined["Current Premium/Barrel"].apply(
                        lambda x: f"{x:.3f}" if pd.notna(x) else "0.000"
                    )
                
                st.dataframe(display_combined, use_container_width=True, height=600)
                st.info(f"Total trades: {len(display_combined):,} | Calculated USD columns: {len(usd_columns)}")
                
                # Download option for combined data
                output_combined = BytesIO()
                with pd.ExcelWriter(output_combined, engine='openpyxl') as writer:
                    combined_df.to_excel(writer, sheet_name='Trade Details with Pricing', index=False)
                    
                    # Format the worksheet
                    worksheet = writer.sheets['Trade Details with Pricing']
                    format_excel_worksheet(worksheet, combined_df)
                
                output_combined.seek(0)
                
                st.download_button(
                    label="📥 Download Complete Trade Details with Pricing (Excel)",
                    data=output_combined,
                    file_name=f"trade_details_with_pricing_{selected_portfolio}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("Please select a portfolio from the sidebar")
    
    # Tab 6: ITM Analysis
    with tab6:
        st.header(f"ITM Analysis - {selected_portfolio}")
        
        if selected_portfolio:
            # Filter by selected portfolio
            itm_df = df[df["PortfolioName"] == selected_portfolio].copy()
            itm_df["BuyOrSell"] = pd.to_numeric(itm_df["BuyOrSell"], errors="coerce")
            
            itm_df["StartFixDate"] = pd.to_datetime(itm_df["StartFixDate"])
            itm_df["EndFixDate_MaturityDate"] = pd.to_datetime(itm_df["EndFixDate_MaturityDate"])
            
            # Create monthly breakdown for ALL trades (both buy and sell)
            monthly_rows = []
            for _, row in itm_df.iterrows():
                start = row["StartFixDate"].replace(day=1)
                end = row["EndFixDate_MaturityDate"].replace(day=1)
                months_diff = (end.year - start.year) * 12 + (end.month - start.month) + 1
                monthly_quantity = row["Position_Quantity"] / months_diff
                
                # Determine leg type and strike price
                is_buy_leg = row["BuyOrSell"] == 1
                
                if is_buy_leg:
                    # Buy leg: StrikePrice1 for protection, StrikePrice2 for sell side
                    buy_strike = row["StrikePrice1"]
                    sell_strike = row["StrikePrice2"]
                else:
                    # Sell leg (BuyOrSell = NaN/None/0): Only StrikePrice1
                    buy_strike = None
                    sell_strike = row["StrikePrice1"]
                
                for m in range(months_diff):
                    month = start + pd.DateOffset(months=m)
                    
                    # Add buy leg entry if applicable
                    if buy_strike is not None and pd.notna(buy_strike):
                        monthly_rows.append({
                            "Month": month.strftime('%Y-%m'),
                            "StrikePrice": buy_strike,
                            "Leg_Type": "Buy Leg",
                            "Monthly_Position_Quantity": monthly_quantity
                        })
                    
                    # Add sell leg entry if applicable
                    if sell_strike is not None and pd.notna(sell_strike):
                        monthly_rows.append({
                            "Month": month.strftime('%Y-%m'),
                            "StrikePrice": sell_strike,
                            "Leg_Type": "Sell Leg",
                            "Monthly_Position_Quantity": monthly_quantity
                        })
            
            monthly_itm_df = pd.DataFrame(monthly_rows)
            
            if not monthly_itm_df.empty:
                # DataFrame 1: Buy Leg (BuyOrSell = 1) - StrikePrice1
                st.subheader("Buy Leg (Protection) - Position Quantity by Month")
                
                buy_leg_df = monthly_itm_df[monthly_itm_df["Leg_Type"] == "Buy Leg"]
                
                if not buy_leg_df.empty:
                    pivot_buy = buy_leg_df.pivot_table(
                        index="StrikePrice",
                        columns="Month",
                        values="Monthly_Position_Quantity",
                        aggfunc="sum",
                        fill_value=0
                    )
                    
                    # Sort columns chronologically
                    pivot_buy = pivot_buy[sorted(pivot_buy.columns)]
                    
                    # Format numbers
                    formatted_buy = pivot_buy.applymap(lambda x: f"{x:,.0f}" if x != 0 else "0")
                    
                    st.dataframe(formatted_buy, use_container_width=True)
                else:
                    st.info("No buy leg trades found for this portfolio")
                
                # DataFrame 2: Sell Leg (BuyOrSell = None/0) - StrikePrice2
                st.subheader("Sell Leg - Position Quantity by Month")
                
                sell_leg_df = monthly_itm_df[monthly_itm_df["Leg_Type"] == "Sell Leg"]
                
                if not sell_leg_df.empty:
                    pivot_sell = sell_leg_df.pivot_table(
                        index="StrikePrice",
                        columns="Month",
                        values="Monthly_Position_Quantity",
                        aggfunc="sum",
                        fill_value=0
                    )
                    
                    # Sort columns chronologically
                    pivot_sell = pivot_sell[sorted(pivot_sell.columns)]
                    
                    # Format numbers
                    formatted_sell = pivot_sell.applymap(lambda x: f"{x:,.0f}" if x != 0 else "0")
                    
                    st.dataframe(formatted_sell, use_container_width=True)
                else:
                    st.info("No sell leg trades found for this portfolio")
                
                # DataFrame 3: Monthly Average Brent Price
                st.subheader("Monthly Average Brent Price (Editable)")
                
                # Get all years from the position data
                all_months_in_data = sorted(monthly_itm_df["Month"].unique())
                
                if all_months_in_data:
                    # Extract years from position data
                    years_in_data = sorted(set([int(m.split('-')[0]) for m in all_months_in_data]))
                    
                    # Always start from 2021 and include all years up to max year in data
                    start_year = 2021
                    end_year = max(years_in_data)
                    all_years = list(range(start_year, end_year + 1))
                    
                    # Fetch historical Brent data
                    start_date = pd.Timestamp(f"{start_year}-01-01")
                    end_date = pd.Timestamp.today()
                    
                    with st.spinner("Fetching Brent crude oil prices..."):
                        brent = load_brent_data(start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))
                        
                        # Create empty dataframe with all years
                        month_order = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", 
                                    "Aug", "Sep", "Oct", "Nov", "Dec"]
                        
                        # Initialize with empty dataframe
                        brent_pivot = pd.DataFrame(index=all_years, columns=month_order)
                        
                        if not brent.empty:
                            # Convert to datetime
                            brent["Date"] = pd.to_datetime(brent["Date"])
                            
                            # Extract year and month names
                            brent["Year"] = brent["Date"].dt.year
                            brent["Month"] = brent["Date"].dt.strftime('%b')
                            
                            # Calculate monthly averages by year
                            monthly_avg = brent.groupby(["Year", "Month"])["Close"].mean().reset_index()
                            
                            # Pivot so each row = year, each column = month
                            brent_pivot_temp = monthly_avg.pivot(index="Year", columns="Month", values="Close")
                            
                            # Update the main pivot with actual data where available
                            for year in brent_pivot_temp.index:
                                if year in brent_pivot.index:
                                    for month in month_order:
                                        if month in brent_pivot_temp.columns:
                                            brent_pivot.at[year, month] = brent_pivot_temp.at[year, month]
                        
                        # Set index name
                        brent_pivot.index.name = "Year"
                        
                        # Convert to float and handle NaN
                        brent_pivot = brent_pivot.astype(float)
                        
                        # Display editable dataframe
                        st.info("💡 You can edit the cells below to input future prices or adjust historical prices")
                        
                        edited_brent = st.data_editor(
                            brent_pivot,
                            use_container_width=True,
                            num_rows="fixed",
                            column_config={
                                month: st.column_config.NumberColumn(
                                    month,
                                    format="$%.2f",
                                    min_value=0,
                                    max_value=500
                                ) for month in month_order
                            }
                        )
                        
                        # Option to download edited Brent prices
                        if edited_brent is not None:
                            brent_csv = edited_brent.to_csv()
                            st.download_button(
                                label="📥 Download Brent Prices (CSV)",
                                data=brent_csv,
                                file_name=f"brent_prices_{selected_portfolio}.csv",
                                mime="text/csv"
                            )
                            
                            st.markdown("---")
                            
                            # Button to calculate payoffs
                            if st.button("🧮 Calculate Payoffs", type="primary"):
                                # Convert edited_brent to month-year format for matching
                                brent_monthly_dict = {}
                                month_to_num = {
                                    "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
                                    "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
                                    "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12"
                                }
                                
                                for year in edited_brent.index:
                                    for month_name in month_order:
                                        month_num = month_to_num[month_name]
                                        month_str = f"{year}-{month_num}"
                                        brent_price = edited_brent.at[year, month_name]
                                        if pd.notna(brent_price):
                                            brent_monthly_dict[month_str] = float(brent_price)
                                
                                # DataFrame 4: Buy Leg Payoff
                                st.subheader("Buy Leg Payoff - max(Brent - Strike, 0) × Position")
                                
                                buy_leg_df = monthly_itm_df[monthly_itm_df["Leg_Type"] == "Buy Leg"]
                                
                                if not buy_leg_df.empty:
                                    # Create payoff matrix
                                    buy_payoff_rows = []
                                    
                                    for strike in sorted(buy_leg_df["StrikePrice"].unique()):
                                        row_data = {"Strike Price": strike}
                                        
                                        for month in sorted(buy_leg_df["Month"].unique()):
                                            # Get position for this strike and month
                                            position = buy_leg_df[
                                                (buy_leg_df["StrikePrice"] == strike) & 
                                                (buy_leg_df["Month"] == month)
                                            ]["Monthly_Position_Quantity"].sum()
                                            
                                            # Get Brent price for this month
                                            brent_price = brent_monthly_dict.get(month, None)
                                            
                                            # Calculate payoff: max(Brent - Strike, 0) × Position
                                            if brent_price is not None and position > 0:
                                                payoff = max(strike - brent_price , 0) * position
                                            else:
                                                payoff = 0
                                            
                                            row_data[month] = payoff
                                        
                                        buy_payoff_rows.append(row_data)
                                    
                                    buy_payoff_df = pd.DataFrame(buy_payoff_rows)
                                    buy_payoff_df = buy_payoff_df.set_index("Strike Price")
                                    
                                    # Add Brent Price row
                                    brent_price_row = {}
                                    for month in sorted(buy_leg_df["Month"].unique()):
                                        brent_price = brent_monthly_dict.get(month, None)
                                        brent_price_row[month] = brent_price if brent_price is not None else 0
                                    brent_price_series = pd.Series(brent_price_row, name="Brent Price")
                                    
                                    # Add Total row
                                    total_row = buy_payoff_df.sum()
                                    total_row.name = "TOTAL"
                                    
                                    # Combine: Brent Price, strikes, then TOTAL
                                    buy_payoff_df = pd.concat([
                                        brent_price_series.to_frame().T,
                                        buy_payoff_df, 
                                        total_row.to_frame().T
                                    ])
                                    
                                    # Format numbers - different format for Brent Price row
                                    formatted_buy_payoff = buy_payoff_df.copy()
                                    for col in formatted_buy_payoff.columns:
                                        for idx in formatted_buy_payoff.index:
                                            val = formatted_buy_payoff.at[idx, col]
                                            if idx == "Brent Price":
                                                # Format Brent prices with $ and 2 decimals
                                                formatted_buy_payoff.at[idx, col] = f"${val:.2f}" if pd.notna(val) and val != 0 else "-"
                                            else:
                                                # Format payoffs with $ and comma
                                                formatted_buy_payoff.at[idx, col] = f"${val:,.2f}" if pd.notna(val) and val != 0 else "$0.00"
                                    
                                    st.dataframe(formatted_buy_payoff, use_container_width=True)
                                    
                                    # Add total metric (exclude Brent Price and TOTAL rows)
                                    total_buy_payoff = buy_payoff_df.iloc[1:-1].sum().sum()
                                    st.metric("Total Buy Leg Payoff", f"${total_buy_payoff:,.2f}")
                                else:
                                    st.info("No buy leg trades found for payoff calculation")
                                
                                # DataFrame 5: Sell Leg Payoff
                                st.subheader("Sell Leg Payoff - max(Strike - Brent, 0) × Position")
                                
                                sell_leg_df = monthly_itm_df[monthly_itm_df["Leg_Type"] == "Sell Leg"]
                                
                                if not sell_leg_df.empty:
                                    # Create payoff matrix
                                    sell_payoff_rows = []
                                    
                                    for strike in sorted(sell_leg_df["StrikePrice"].unique()):
                                        row_data = {"Strike Price": strike}
                                        
                                        for month in sorted(sell_leg_df["Month"].unique()):
                                            # Get position for this strike and month
                                            position = sell_leg_df[
                                                (sell_leg_df["StrikePrice"] == strike) & 
                                                (sell_leg_df["Month"] == month)
                                            ]["Monthly_Position_Quantity"].sum()
                                            
                                            # Get Brent price for this month
                                            brent_price = brent_monthly_dict.get(month, None)
                                            
                                            # Calculate payoff: max(Brent - Strike, 0) × Position
                                            if brent_price is not None and position > 0:
                                                payoff = max(strike - brent_price, 0) * position
                                            else:
                                                payoff = 0
                                            
                                            row_data[month] = payoff
                                        
                                        sell_payoff_rows.append(row_data)
                                    
                                    sell_payoff_df = pd.DataFrame(sell_payoff_rows)
                                    sell_payoff_df = sell_payoff_df.set_index("Strike Price")
                                    
                                    # Add Brent Price row
                                    brent_price_row = {}
                                    for month in sorted(sell_leg_df["Month"].unique()):
                                        brent_price = brent_monthly_dict.get(month, None)
                                        brent_price_row[month] = brent_price if brent_price is not None else 0
                                    brent_price_series = pd.Series(brent_price_row, name="Brent Price")
                                    
                                    # Add Total row
                                    total_row = sell_payoff_df.sum()
                                    total_row.name = "TOTAL"
                                    
                                    # Combine: Brent Price, strikes, then TOTAL
                                    sell_payoff_df = pd.concat([
                                        brent_price_series.to_frame().T,
                                        sell_payoff_df, 
                                        total_row.to_frame().T
                                    ])
                                    
                                    # Format numbers - different format for Brent Price row
                                    formatted_sell_payoff = sell_payoff_df.copy()
                                    for col in formatted_sell_payoff.columns:
                                        for idx in formatted_sell_payoff.index:
                                            val = formatted_sell_payoff.at[idx, col]
                                            if idx == "Brent Price":
                                                # Format Brent prices with $ and 2 decimals
                                                formatted_sell_payoff.at[idx, col] = f"${val:.2f}" if pd.notna(val) and val != 0 else "-"
                                            else:
                                                # Format payoffs with $ and comma
                                                formatted_sell_payoff.at[idx, col] = f"${val:,.2f}" if pd.notna(val) and val != 0 else "$0.00"
                                    
                                    st.dataframe(formatted_sell_payoff, use_container_width=True)
                                    
                                    # Add total metric (exclude Brent Price and TOTAL rows)
                                    total_sell_payoff = sell_payoff_df.iloc[1:-1].sum().sum()
                                    st.metric("Total Sell Leg Payoff", f"${total_sell_payoff:,.2f}")
                                else:
                                    st.info("No sell leg trades found for payoff calculation")
                                
                                # DataFrame 6: Net Payoff Summary
                                st.markdown("---")
                                st.subheader("Net Payoff Summary")
                                
                                # Get all unique months from both legs
                                all_months_payoff = sorted(set(
                                    list(buy_leg_df["Month"].unique() if not buy_leg_df.empty else []) +
                                    list(sell_leg_df["Month"].unique() if not sell_leg_df.empty else [])
                                ))
                                
                                if all_months_payoff:
                                    summary_data = {
                                        "Total Buy Leg": {},
                                        "Total Sell Leg": {},
                                        "Net Payoff": {}
                                    }
                                    
                                    for month in all_months_payoff:
                                        # Get buy leg payoff for this month (exclude Brent Price row)
                                        buy_payoff_month = 0
                                        if not buy_leg_df.empty and month in buy_payoff_df.columns:
                                            buy_payoff_month = buy_payoff_df.iloc[1:-1][month].sum()  # Exclude Brent Price and TOTAL rows
                                        
                                        # Get sell leg payoff for this month (exclude Brent Price row)
                                        sell_payoff_month = 0
                                        if not sell_leg_df.empty and month in sell_payoff_df.columns:
                                            sell_payoff_month = sell_payoff_df.iloc[1:-1][month].sum()  # Exclude Brent Price and TOTAL rows
                                        
                                        # Calculate net payoff
                                        net_payoff_month = buy_payoff_month - sell_payoff_month
                                        
                                        summary_data["Total Buy Leg"][month] = buy_payoff_month
                                        summary_data["Total Sell Leg"][month] = sell_payoff_month
                                        summary_data["Net Payoff"][month] = net_payoff_month
                                    
                                    # Create summary dataframe
                                    summary_df = pd.DataFrame(summary_data).T
                                    
                                    # Format numbers
                                    formatted_summary = summary_df.copy()
                                    for col in formatted_summary.columns:
                                        for idx in formatted_summary.index:
                                            val = formatted_summary.at[idx, col]
                                            formatted_summary.at[idx, col] = f"${val:,.2f}" if pd.notna(val) else "$0.00"
                                    
                                    st.dataframe(formatted_summary, use_container_width=True)
                                    
                                    # Display total net payoff metric
                                    total_net_payoff = summary_df.loc["Net Payoff"].sum()
                                    st.metric("Total Net Payoff", f"${total_net_payoff:,.2f}")
                                else:
                                    st.info("No payoff data available for summary")
                else:
                    st.warning("No position data available to determine years")
            else:
                st.warning("No data available for ITM analysis")
        else:
            st.warning("Please select a portfolio from the sidebar")
    
    # Tab 7: Raw Data
    with tab7:
        st.header(f"Raw Data View - {selected_portfolio}")
        
        if selected_portfolio:
            # Filter by selected portfolio
            filtered_df = df[df["PortfolioName"] == selected_portfolio].copy()
            st.dataframe(filtered_df, use_container_width=True, height=600)
            st.info(f"Total rows for {selected_portfolio}: {len(filtered_df):,}")
        else:
            st.dataframe(df, use_container_width=True, height=600)
            st.info(f"Total rows: {len(df):,}")
    
    # Tab 8: Export
    with tab8:
        st.header("Export Data")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Download Cleaned Data")
            csv = df.to_csv(index=False)
            st.download_button(
                label="📥 Download CSV",
                data=csv,
                file_name="portfolio_data_cleaned.csv",
                mime="text/csv"
            )
        
        with col2:
            st.subheader("Download Summary")
            summary_csv = summary.to_csv(index=False)
            st.download_button(
                label="📥 Download Summary CSV",
                data=summary_csv,
                file_name="portfolio_summary.csv",
                mime="text/csv"
            )

else:
    st.info("⏳ Loading data...")
    
    # Show example data structure
    with st.expander("ℹ️ Expected Data Format"):
        st.markdown("""
        The API should return data with the following columns:
        - `TradeNumber`
        - `PortfolioName`
        - `DealerID`
        - `Acronym`
        - `Position_Quantity`
        - `BuyOrSell`
        - `NetPremium`
        - `CashflowUSD`
        - `AbsNetPremium`
        - `StrikePrice1`
        - `StrikePrice2`
        - `StartFixDate`
        - `EndFixDate_MaturityDate`
        """)