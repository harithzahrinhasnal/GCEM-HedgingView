import pandas as pd
from datetime import date
import datetime
import openpyxl
import os
import pyautogui
from AppOpener import open
import time
import pyperclip
from tabulate import tabulate  
from threading import Thread
from collections.abc import Mapping
import openpyxl
import subprocess
import re
import ctypes 
import pymsteams
import pandas as pd
from itertools import product

Portfolio = "FY2025 PCHP"

#Time:
current_date_time = datetime.datetime.now()
formatted_date_time = current_date_time.strftime('%d-%m-%Y')

# Replace 'your_excel_file.xlsx' with the path to your Excel file
excel_file = pd.ExcelFile('PCHP Data.xlsx')

# Assuming the sheet name is 'Overall_data', you can change it accordingly
sheet_name = 'Overall_data'

# Read the Excel sheet into a DataFrame
df = pd.read_excel(excel_file, sheet_name=sheet_name)

# Specify the headers
headers = ['O.January', 'O.February', 'O.March', 'O.April', 'O.May', 'O.June', 'O.July', 'O.August', 'O.September', 'O.October', 'O.November', 'O.December']

# Initialize a list to store columns with values
columns_with_values = []

df = df[df['Portfolio'] == Portfolio]
print(df)

# Iterate through each header and check if it has any non-null values
for header in headers:
    if df[header].notnull().any():
        columns_with_values.append(header)




# Filter DataFrame to include only rows where at least one of the selected columns has a non-null value
filtered_df = df[df[columns_with_values].notnull().any(axis=1)]

# Get unique values of FO.StrikePrice1 and FO.StrikePrice2
unique_values_FO_StrikePrice1 = filtered_df['FO.StrikePrice1'].dropna().astype(int).unique()
unique_values_FO_StrikePrice2 = filtered_df['FO.StrikePrice2'].dropna().astype(int).unique()

# Print unique values
print("Months with Unmatured Volumes:", columns_with_values)
print("Unique values in FO.StrikePrice1:", unique_values_FO_StrikePrice1)
print("Unique values in FO.StrikePrice2:", unique_values_FO_StrikePrice2)



################################################################
current_date = date.today()
current_year = current_date.year
current_month = current_date.month

# Initialize a list to store the updated months
updated_months = []

# Iterate over the original months
for i, month_str in enumerate(columns_with_values, start=1):
    # Extract the month from the column name
    month_name = month_str.split('.')[1]

    # Convert month name to a numerical value
    months_dict = {
        'January': 1, 'February': 2, 'March': 3, 'April': 4,
        'May': 5, 'June': 6, 'July': 7, 'August': 8,
        'September': 9, 'October': 10, 'November': 11, 'December': 12
    }
    month_value = months_dict[month_name]

    # Determine future month and year
    if current_month > month_value:
        future_month = month_value
        future_year = current_year + 1
    else:
        future_month = month_value
        future_year = current_year

    # Format the month and year strings as 'MMYY'
    month_string = f"{future_month:02d}01{future_year % 100:02d}"

    # Append the formatted string to the updated_months list
    updated_months.append(month_string)

# Get unique strike values
unique_strike_values = pd.concat([filtered_df['FO.StrikePrice1'], filtered_df['FO.StrikePrice2']]).dropna().astype(int).unique()

##############
def main():
    x =+ 1
    print('running the code ')
    #Bloomberg Initialization
    def open_bloomberg():
        subprocess.run(['start', '', 'C:/blp/Wintrv/wintrv.exe'], shell=True)
        time.sleep(1)

    open_bloomberg()


    #Choose Your Monitor and Location (Use Coordinate.py to recalibration position of option calc in BBG)
    Monitor = 'BBG Laptop' #BBG Laptop, Probook, Office Monitor
    #LOCATIONS
    if Monitor == 'BBG Laptop': #make sure resolution is (1280,720)
        search_bar = [44, 112]
        CO_1 = [81, 300]
        date_input = [220, 320]
        period = [330 ,330]
        delivery = [280, 340]
        barrel_1 = [280, 480]
        premium_1 = [780, 480]
        calc_time = [361 ,270]
        Upper_strike = [570, 480]
        calc = [82, 181] 
        swap = [399,580]

    if Monitor == 'Probook': #make sure resolution is (1280,720)
        search_bar = [44, 112]
        CO_1 = [81, 300]
        date_input = [220, 310]
        period = [330 ,310]
        delivery = [280, 340]
        barrel_1 = [280, 470]
        premium_1 = [780, 470]
        calc_time = [361 ,250]
        Upper_strike = [570, 470]
        calc = [82, 181] 
        swap = [390,580]

        
    if Monitor == 'Office Monitor': 
        search_bar = [44, 112]
        CO_1 = [81, 300]
        date_input = [220, 310]
        today_input = [220, 250]
        period = [330 ,310]
        delivery = [280, 340]
        barrel_1 = [280, 470]
        premium_1 = [780, 470]
        calc_time = [361 ,250]
        Upper_strike = [570, 470]
        calc = [82, 181]
        swap = [375,635]
    ##############################
    # Create a folder with the name formatted_date_time if it doesn't exist
    folder_name = str(Portfolio)+"_"+formatted_date_time 
    screenshot_folder = 'SCREENSHOT'
    folder_path = os.path.join(os.getcwd(), screenshot_folder, folder_name)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)


    def take_screenshot(x):
        # Get the screen resolution
        screen_width, screen_height = pyautogui.size()

        # Take a screenshot of the entire screen
        screenshot = pyautogui.screenshot()

        # Save the screenshot to a file with the specified name
        screenshot_path = os.path.join(folder_path, f'screenshot_{x}.png')
        screenshot.save(screenshot_path)

        print(f"Screenshot taken and saved as 'screenshot_{x}.png'.")


    def checklevel(IN):
        if IN == '0' or not re.search(r"\d+\.\d+", IN):
            return 0
        numbers = re.findall(r"\d+\.\d+", IN)
        return float(numbers[0])

    time.sleep(1)
    #pyautogui.hotkey('ctrl', 't')

    """
    time.sleep(2)
    pyautogui.moveTo(search_bar, duration = 0.5)
    pyautogui.click()
    time.sleep(2)
    
    pyautogui.write('Co1 Comdty')
    time.sleep(2)
    pyautogui.moveTo(CO_1, duration = 0.5)
    pyautogui.click()"""

    ###OPEN OPTION PRICING
    time.sleep(2)
    #pyautogui.moveTo(search_bar, duration = 0.5)
   # pyautogui.press('enter')
   # pyautogui.write('OVML PT APO')
    #pyautogui.press('enter')
    #time.sleep(9)

    ###OPTION PRICING SETUP####
    pyautogui.moveTo(delivery, duration = 0.5)
    pyautogui.click()
    pyautogui.write('14')

    
    #pyautogui.moveTo(today_input, duration = 0.5)
    #pyautogui.click()


    pyautogui.moveTo(barrel_1, duration = 0.5)
    pyautogui.click()
    pyautogui.write('1')

    pyautogui.moveTo(period, duration = 0.5)
    pyautogui.click()
    pyautogui.write('1')

    pyautogui.moveTo(calc_time, duration = 0.5)
    pyautogui.click()
    pyautogui.write('06:00')

    def get_premium_data(Month, Strike):
        
        Premium = [[strike, 'BBG value'] for strike in Strike]

        ###### OPTION ########
        pyautogui.moveTo(date_input, duration = 0.5)
        pyautogui.click()
        pyautogui.write(Month)
        print(Month)

        for i,k in enumerate(Strike):
            pyautogui.moveTo(Upper_strike, duration = 0.5)
            pyautogui.click()
            pyautogui.write(str(k))

            pyautogui.moveTo(calc, duration = 0.5)
            pyautogui.click()

            if i == 0:
                time.sleep(10)
            else:
                time.sleep(2)

            pyautogui.moveTo(premium_1, duration = 0.5)  
            pyautogui.click()
            pyautogui.click()
            pyautogui.hotkey('ctrl', 'c')


            Premium[i][1] = checklevel (pyperclip.paste())
            take_screenshot( str(Portfolio) +"_"+ str(Month)+'_'+str(k))
            time.sleep(2)
        return Premium

    # Path to the Excel file
    excel_file_path = r'BBG_Output.xlsx'

    # Check if the file exists
    if os.path.exists(excel_file_path):
        # File exists, load the workbook
        wb = openpyxl.load_workbook(excel_file_path)
    else:
        # File doesn't exist, create a new workbook
        wb = openpyxl.Workbook()
        # Save the workbook to the specified path
        wb.save(excel_file_path)


    Strike = unique_strike_values

    
    # Get the sheet title from Day_data
    sheet_title = str(Portfolio) + '_' +str(formatted_date_time) 

    # Check if a sheet with the same title already exists
    if sheet_title in wb.sheetnames:
        # Delete the existing sheet
        sheet_to_delete = wb[sheet_title]
        wb.remove(sheet_to_delete)

    # Create a new worksheet for the current month and set the title to the month value
    ws = wb.create_sheet(title=sheet_title)

    # Set the headers for the table
    headers = [''] + [str(strike) for strike in unique_strike_values] + ['Swap Ref']
    ws.append(headers)

    # Iterate over the months and add a row for each month
    for month in columns_with_values:
        row_data = [month] + [''] * len(Strike)  # Initialize a row with month and empty values
        ws.append(row_data)


    #start taking data from BBG
    for i, m in enumerate(updated_months):
        Premium= get_premium_data( m ,Strike)
        pyautogui.vscroll(-100)
        time.sleep(1)

        pyautogui.moveTo(swap, duration = 0.5)
        pyautogui.click()
        pyautogui.click()
        time.sleep(0.5)
        pyautogui.hotkey('ctrl', 'c')
        ref = float(pyperclip.paste())
        print(ref)
        time.sleep(0.5)
        pyautogui.vscroll(100)


        # Write the premium data to the worksheet
        for strike_index, strike_value in enumerate(Strike, start=1):
            # Get the row index corresponding to the strike
            col_num = strike_index + 1
            row_num =i + 2
            # Get the premium value for the current strike and month
            premium_value = Premium[strike_index - 1][1]  # Index adjusted for 0-based index
            # Write the premium value in the corresponding cell
            ws.cell(row=row_num, column=col_num, value=premium_value)
            ws.cell(row=row_num, column=len(Strike) + 2, value=ref)


    # Save the workbook to a file
    wb.save(r'BBG_Output.xlsx')
    wb.close()

    # Load the Excel file
    file_path = 'BBG_output.xlsx'
    sheet_name = sheet_title  # Assuming the sheet name is '06-02-2024'
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Load workbook and active sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Transpose data
    transposed_data = list(zip(*sheet.iter_rows(values_only=True)))

    # Clear existing data in the sheet
    sheet.delete_rows(1, sheet.max_row)

    # Write transposed data to the sheet
    for row_data in transposed_data:
        sheet.append(row_data)

    # Save changes
    workbook.save(file_path)

    print("Data transposed and saved successfully to the same sheet.")
    ctypes.windll.user32.MessageBoxW(0, "Bot is Done!!", "Pricing Bot",  0x1000)

if __name__ == '__main__':
    main()





